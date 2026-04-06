"""FileReaderAgent — lee y extrae contenido de archivos Excel, PDF, Word, CSV, TXT."""
from __future__ import annotations
from pathlib import Path
from core.base_agent import BaseAgent
from core.context import AgentContext


class FileReaderAgent(BaseAgent):
    name = "FileReaderAgent"
    description = "Lee y extrae contenido estructurado de archivos Excel, PDF, Word, CSV o TXT."

    async def run(self, context: AgentContext) -> AgentContext:
        file_path = getattr(context, 'input_file', None)
        if not file_path or not Path(file_path).exists():
            self.log(context, "No se especifico archivo de entrada o no existe")
            context.set_data('file_content', context.user_input)
            context.set_data('file_meta', {'type': 'text', 'name': 'user_input', 'size': len(context.user_input)})
            return context

        ext = Path(file_path).suffix.lower()
        file_name = Path(file_path).name
        self.log(context, f"Leyendo {ext}: {file_name}")

        content = self._read_file(file_path, ext)
        meta = {
            'name': file_name,
            'type': ext,
            'size_bytes': Path(file_path).stat().st_size,
            'chars_extracted': len(content),
        }

        context.set_data('file_content', content)
        context.set_data('file_meta', meta)
        self.log(context, f"Extraidos {len(content)} chars de {file_name}")
        return context

    def _read_file(self, file_path: str, ext: str) -> str:
        try:
            if ext == '.csv':
                import csv
                rows = []
                with open(file_path, encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    for i, row in enumerate(reader):
                        rows.append(str(dict(row)))
                        if i > 100:
                            rows.append('... (truncado a 100 filas)')
                            break
                return '\n'.join(rows)
            elif ext in ('.xlsx', '.xls'):
                import openpyxl
                wb = openpyxl.load_workbook(file_path, data_only=True)
                result = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    result.append(f"=== HOJA: {sheet} ===")
                    for row in list(ws.iter_rows(values_only=True))[:50]:
                        result.append(str(row))
                return '\n'.join(result)
            elif ext == '.pdf':
                import pdfplumber
                with pdfplumber.open(file_path) as pdf:
                    return '\n'.join(p.extract_text() or '' for p in pdf.pages[:20])
            elif ext in ('.docx', '.doc'):
                import docx
                doc = docx.Document(file_path)
                return '\n'.join(p.text for p in doc.paragraphs if p.text.strip())
            elif ext in ('.txt', '.md', '.py', '.json'):
                return Path(file_path).read_text(encoding='utf-8')
        except ImportError as e:
            return f"Falta libreria: {e}. Instala con: pip install openpyxl pdfplumber python-docx"
        except Exception as e:
            return f"Error leyendo archivo: {e}"
        return ''
