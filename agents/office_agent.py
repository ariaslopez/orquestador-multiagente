"""OfficeAgent — Lee y analiza archivos Office (Excel, PDF, Word, CSV)."""
from __future__ import annotations
from pathlib import Path
from core.base_agent import BaseAgent
from core.context import AgentContext


class OfficeAgent(BaseAgent):
    name = "OfficeAgent"
    description = "Lee y extrae datos de archivos Excel, PDF, Word, CSV para analisis."

    async def run(self, context: AgentContext) -> AgentContext:
        file_path = getattr(context, 'input_file', None)
        if not file_path or not Path(file_path).exists():
            self.log(context, "No se especifico archivo de entrada o no existe")
            return context

        ext = Path(file_path).suffix.lower()
        self.log(context, f"Leyendo archivo {ext}: {file_path}")
        content = self._read_file(file_path, ext)

        if not content:
            self.log(context, "No se pudo extraer contenido del archivo")
            return context

        prompt = f"""Analiza el siguiente contenido extraido del archivo `{Path(file_path).name}`.

PETICION DEL USUARIO: {context.user_input}

CONTENIDO DEL ARCHIVO:
{content[:4000]}

Genera:
1. RESUMEN EJECUTIVO (hallazgos mas importantes)
2. METRICAS CLAVE (datos numericos relevantes)
3. PATRONES DETECTADOS (tendencias o anomalias)
4. RECOMENDACIONES DE ACCION (proximos pasos concretos)
5. DATOS PARA EXPORTAR (tabla resumen en formato markdown)"""
        analysis = await self.llm(context, prompt, temperature=0.2)
        context.final_output = analysis
        context.pipeline_name = "OFFICE"
        self.log(context, "Analisis de archivo completado")
        return context

    def _read_file(self, file_path: str, ext: str) -> str:
        try:
            if ext == '.csv':
                import csv
                rows = []
                with open(file_path, encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f)
                    for i, row in enumerate(reader):
                        rows.append(str(dict(row)))
                        if i > 100:
                            rows.append('... (truncado a 100 filas)')
                            break
                return '\n'.join(rows)
            elif ext in ('.xlsx', '.xls'):
                import openpyxl
                wb = openpyxl.load_workbook(file_path, data_only=True)
                result = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    result.append(f"=== HOJA: {sheet} ===")
                    for row in list(ws.iter_rows(values_only=True))[:50]:
                        result.append(str(row))
                return '\n'.join(result)
            elif ext == '.pdf':
                import pdfplumber
                with pdfplumber.open(file_path) as pdf:
                    return '\n'.join(p.extract_text() or '' for p in pdf.pages[:20])
            elif ext in ('.docx', '.doc'):
                import docx
                doc = docx.Document(file_path)
                return '\n'.join(p.text for p in doc.paragraphs if p.text.strip())
            elif ext == '.txt':
                return Path(file_path).read_text(encoding='utf-8')
        except ImportError as e:
            return f"Falta libreria: {e}. Instala con: pip install openpyxl pdfplumber python-docx"
        except Exception as e:
            return f"Error leyendo archivo: {e}"
        return ''
